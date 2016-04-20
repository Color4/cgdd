from django.http import HttpResponse
from django.shortcuts import get_object_or_404, render

#Is this image comple as was locked by R:
#    SEMG2_CAMK1_PANCAN__PMID26947069.png
 
 
from .models import Study, Gene, Dependency  # Removed: Histotype,
import json # For ajax for the jquery autocomplete search box
import math # For ceil()
from datetime import datetime # For get_timming()

# This django logging is configured in settings.py and is based on: http://ianalexandr.com/blog/getting-started-with-django-logging-in-5-minutes.html
#import logging
#logger = logging.getLogger(__name__)
#def log(): logger.debug("this is a debug message!")
#def log_error(): logger.error("this is an error message!!")

json_mimetype = 'application/json; charset=utf-8'
html_mimetype = 'text/html; charset=utf-8'
    
def json_error(message, status_code='0'):    
    return HttpResponse( json.dumps( {'success': False, 'error': status_code, 'message': message } ), json_mimetype ) # eg: str(exception)

def is_search_by_driver(search_by):
    if   search_by == 'driver': return True
    elif search_by == 'target': return False
    else: print("ERROR: **** Invalid search_by: '%s' ****" %(search_by))

    
    
def get_timing(start_time, name, time_array=None):
    # To print timings, and optionally build an array of timings that can then be sent to webbrowser console via json.
    # The start_time should be obtained from: datetime.now()
    duration = datetime.now() - start_time
    print( "%s: %s msec" %(name,str(duration)))  # or use: duration.total_seconds()
    if time_array is not None:
        #if name in time_dict: print("WARNING: Key '%s' is already in the time_dict" %(name))
        time_array.append({name: str(duration)}) # Uses an array so will preserve the order, rather than a dictionary.
    return datetime.now()


def index(request, search_by = 'driver', gene_name=''): # Default is search boxes, with gene dropdown populated with driver gene_names (plus an empty name).
    
    driver_list = Gene.objects.filter(is_driver=True).only("gene_name", "full_name", "is_driver", "prevname_synonyms").order_by('gene_name')  # Needs: (is_driver=True), not just: (is_driver)

    target_list = []
    if not is_search_by_driver(search_by):
        target_list = Gene.objects.filter(is_target=True).only("gene_name", "full_name", "is_target", "prevname_synonyms").order_by('gene_name')  # Needs: (is_target=True), not just: (is_target)

    # histotype_list = Histotype.objects.order_by('full_name')
    histotype_list = Dependency.HISTOTYPE_CHOICES
    experimenttype_list = Study.EXPERIMENTTYPE_CHOICES
    study_list = Study.objects.order_by('pmid')
    dependency_list = None # For now.
        
    # This page can be called from the 'drivers' or 'targets' page, with a driver as a POST parameter, so then should display the POST results?
    if gene_name == '': # if gene not passed using the '/gene_name' parameter in urls.py
        if   request.method == 'GET':  gene_name = request.GET.get('gene_name', '')
        elif request.method == 'POST': gene_name = request.POST.get('gene_name', '')
        else: gene_name = ''
    
    # current_url = request.get_full_path() # To display the host in title for developing on lcalhost or pythonanywhere server.
    # current_url = request.build_absolute_uri()
    #current_url =  request.META['SERVER_NAME']
    current_url =  request.META['HTTP_HOST']

    # Optionally could add locals() to the context to pass all local variables, eg: return render(request, 'app/page.html', locals())
    context = {'search_by': search_by, 'gene_name': gene_name, 'driver_list': driver_list, 'target_list': target_list,'histotype_list': histotype_list, 'study_list': study_list, 'experimenttype_list': experimenttype_list, 'dependency_list': dependency_list, 'current_url': current_url}
    return render(request, 'gendep/index.html', context)


def get_drivers(request):
    # View for the driver_jquery search box.
    #if request.is_ajax(): # Users can also access this from scripts so not always AJAX
    name_contains = request.GET.get('name', '')  #  jQuery autocomplete sends the query as "term" and it expects back three fields: id, label, and value. It will use those to display the label and then the value to autocomplete each driver gene.
    # driver_list = Gene.objects.filter(is_driver=True).order_by('gene_name')  # Needs: (is_driver=True), not just: (is_driver)
    if name_contains != '':
        drivers = Gene.objects.filter(is_driver=True, gene_name__icontains = name_contains)   # [:20]
    else:
        drivers = Gene.objects.filter(is_driver=True)   # [:20]
    
    # ADD: full_name__icontains = q or prev_names__icontains = q or synonyms__icontains = q
    results = []
    for d in drivers:
        d_json = {}
        d_json['id'] = d.gene_name  # But maybe this needs to be an integer? 
        d_json['label'] = d.gene_name
        d_json['value'] = d.gene_name + ' : ' + d.full_name + ' : ' + d.prevname_synonyms
        results.append(d_json)
    data = json.dumps(results)
        # Alternatively use: return HttpResponse(simplejson.dumps( [drug.field for drug in drugs ]))
        # eg: format is:
        # [ {"id": "3", "value":"3","label":"Matching employee A"},
        #   {"id": "5", "value":"5","label":"Matching employee B"},
        # ]
    # data = json.dumps(list(Town.objects.filter(name__icontains=q).values('name')))
    
    #    data = 'fail'

    return HttpResponse(data, json_mimetype)


def build_dependency_query(search_by, gene_name, histotype_name, study_pmid, wilcox_p=0.05, order_by='wilcox_p', select_related=None):
    error_msg = ""
    
    if gene_name == "":
        error_msg += 'Gene name is empty, but must be specified'
        return error_msg, None, None, None, None

    # As Query Sets are lazy, so can build query and evaluated once at end:
    q = Dependency.objects.filter(wilcox_p__lte = wilcox_p) # Only list significant hits (ie: p<=0.05)
        
    try: # To not need the table join, 
        gene = Gene.objects.get(gene_name=gene_name)
        # gene = None if gene_name == '' else Gene.objects.get(gene_name=gene_name) 
        if is_search_by_driver(search_by):
            q = q.filter( driver = gene )  # q = q.filter( driver__gene_name = gene )  maybe use driver_id = gene_name
        else:
            q = q.filter( target = gene )  # q = q.filter( target__gene_name = gene )  maybe use target_id = gene_name
        
    except ObjectDoesNotExist: # Not found by the objects.get()
        error_msg = "Gene '%s' NOT found in Gene table" %(gene_name)  # if gene is None
        return error_msg, None, None, None, None

    if histotype_name != "ALL_HISTOTYPES":
        histotype_full_name = Dependency.histotype_full_name(histotype_name)
        q = q.filter( histotype = histotype_name ) # This correctly uses "...= histotype_name" (not "...= histotype_full_name")
    else: 
        histotype_full_name = "All tissues"

    if study_pmid != "ALL_STUDIES":
        try:
            study = Study.objects.get(pmid=study_pmid)
            q = q.filter( study = study )
        except ObjectDoesNotExist: # Not found by the objects.get()
            error_msg += " Study pmid='%s' NOT found in Study table" %(study_pmid)
            return error_msg, None, None, None, None
    else:
        study = "ALL_STUDIES"

    if select_related != None:
        if isinstance(select_related, str) and select_related != '': 
            q = q.select_related(select_related)
        elif isinstance(select_related, list):
            for column in select_related:
                q = q.select_related(column)
        else:
            error_msg += " ERROR: *** Invalid type for 'select_related' ***"        
            print(error_msg)

     
    if order_by != None and order_by != '':
        q = q.order_by(order_by)  # usually 'wilcox_p', but was: order_by('target__gene_name')
        
    return error_msg, q, gene, histotype_full_name, study
    
    
def ajax_results_slow_full_detail_version(request):
    # View for the dependency search result table.
    # Ajax sends four fields: driver, histotype, pmid, [start(row for pagination):

    request_method = request.method # 'POST' or 'GET'
    if request_method != 'POST': return HttpResponse('Expected a POST request, but got a %s request' %(request_method), html_mimetype)

    if not request.is_ajax(): return HttpResponse('fail', html_mimetype)

    driver_name = request.POST.get('driver', "")  # It's an ajax POST request, rather than the usual ajax GET request
    histotype_name = request.POST.get('histotype', "ALL_HISTOTYPES")
    study_pmid = request.POST.get('study', "ALL_STUDIES")

    error_msg, dependency_list, driver, histotype_full_name, study = build_dependency_query(driver_name, histotype_name, study_pmid)
    
    if error_msg != '': return HttpResponse("Error: "+error_msg, html_mimetype)

    gene_weblinks = driver.external_links('|')
       
    current_url =  request.META['HTTP_HOST']  # or: request.META['SERVER_NAME']

    context = {'dependency_list': dependency_list, 'driver': driver, 'histotype': histotype_name, 'histotype_full_name': histotype_full_name, 'study': study, 'gene_weblinks': gene_weblinks, 'current_url': current_url}
    return render(request, 'gendep/ajax_results.html', context, content_type=html_mimetype) #  ??? .. charset=utf-8"


def gene_ids_as_dictionary(gene):
  return {
    'gene_name': gene.gene_name,
    'entrez_id': gene.entrez_id,
    'ensembl_id': gene.ensembl_id,
    'ensembl_protein_id': gene.ensembl_protein_id,
    'vega_id': gene.vega_id,
    'omim_id': gene.omim_id,
    'hgnc_id': gene.hgnc_id,
    'cosmic_id': gene.cosmic_id,
    'uniprot_id': gene.uniprot_id
    }
    
    
# The following are cached so don't need to reload on each request:
# These map the study_pmid and histotype short name to an integer:
study_dict = dict()  # Maybe simpler to hard code the studies in 'models.py' in future.
# Then could use small integer in the dependency table for study and histotype.
study_json = None
study_last_reloaded = None
histotype_dict = dict() # As histotypes are hardcoded in 'models.py' then won't change while server running.
histotype_json = None



    
def get_dependencies(request, search_by, gene_name, histotype_name, study_pmid):
    # Results JSON formatted data for the dependency search result table.
    # Ajax on from "Search" button on the "index.html" page sends four fields: search_by, gene_name, histotype, pmid, [start(row for pagination):
    # returns json - should also return the error message as json format.
    # Get request is faster than post, as Post make two http requests, Get makes one, the django parameters are a get.
    # mimetype = 'text/html' # for error messages. was: 'application/json'

    from django.core.cache import cache  # To cache previous results. "To provide thread-safety, a different instance of the cache backend will be returned for each thread."    
    
    timing_array = []  # Using an array to preserve order on output.
    start = datetime.now()
    
    ajax_results_cache_version = '1' # version of the data in the database and of this json format. Increment this on updates that change the db or this json format. See: https://docs.djangoproject.com/en/1.9/topics/cache/#cache-versioning

    search_by_driver = is_search_by_driver(search_by) # otherwise is by target
    if search_by_driver: print("***** search_by_driver is TRUE")
    else: print("***** search_by_driver is FALSE")
    
    # Avoid storing a 'None' on the cache as then can't tell if if cache miss or is value of the key
    cache_key = search_by+'_'+gene_name+'_'+histotype_name+'_'+study_pmid+'_v'+ajax_results_cache_version
    cache_data = cache.get(cache_key, 'not_found')
    if cache_data != 'not_found': 
        start = get_timing(start, 'Retrieved from cache', timing_array)
    #    return HttpResponse(cache_data, json_mimetype, version=ajax_results_cache_version)
   
    #d = end - start  # datetime.timedelta object
    #print d.total_seconds()  

    
    #request_method = request.method # 'POST' or 'GET'
    #if request_method != 'POST': return HttpResponse('Expected a POST request, but got a %s request' %(request_method), json_mimetype)

    # if not request.is_ajax(): return HttpResponse('fail', json_mimetype)
    
    # search_by = request.POST.get('search_by', "")  # It's an ajax POST request, rather than the usual ajax GET request
    # gene_name = request.POST.get('gene', "") 
    # histotype_name = request.POST.get('histotype', "ALL_HISTOTYPES")
    # study_pmid = request.POST.get('study', "ALL_STUDIES")
    # effect_size = request.POST.get('effect_size', "ALL_EFFECT_SIZES")
    # return json_error('Hello') # HttpResponse("Hello", json_mimetype)
    
    if search_by_driver:
        select_related = [ 'target__inhibitors', 'target__ensembl_protein_id' ]
    else:
        select_related = [ 'driver__inhibitors', 'driver__ensembl_protein_id' ]
 
    error_msg, dependency_list, gene, histotype_full_name, study = build_dependency_query(search_by,gene_name, histotype_name, study_pmid, order_by='wilcox_p', select_related=select_related) # can add select related if needed, eg: for target gene prevname_synonyms.
    if error_msg != '': return json_error("Error: "+error_msg)

    # gene_weblinks = gene.external_links('|') # Now in javascript
       
    current_url =  request.META['HTTP_HOST']  # or: request.META['SERVER_NAME']

    start = get_timing(start, 'Query setup', timing_array)
        
    results = []
    csv = ''
    div = ';' # Using semicolon as the div, as comma may be used to separate the inhibitors
    
    # From server-side, a csv file format is probably more efficient as is just commas as separater, rather than repeating the keys multiple times.
    # Maybe even use raw sql for maybe faster query: https://docs.djangoproject.com/en/1.9/topics/db/sql/#passing-parameters-into-raw
    # *** To see the raw SQL, try using:  django.db.connection.queries 

    # In MySQL can directly use:   SELECT CONCAT_WS(',', field1, field2, field3) FROM table; 
    #   http://stackoverflow.com/questions/707473/how-do-you-output-mysql-query-results-in-csv-format-to-the-screen-not-to-a-fil
    #   CONCAT_WS() does not skip empty strings. However, it does skip any NULL values after the separator argument. 
    #   https://ariejan.net/2008/11/27/export-csv-directly-from-mysql/
    # Is there anything similar for sqlite?
    
    # If need to acccess fields in a related table, use: ...select_related('fieldname')  https://www.neustar.biz/blog/optimizing-django-queries
    # Although just works on 1-to-1 relationships, tthen use dict approach.
    # Other ideas: https://blog.mozilla.org/webdev/2011/12/15/django-optimization-story-thousand-times-faster/
    
    # Maybe use: json = serializers.serialize('json', playlists)
    #            return HttpResponse(json, mimetype=json_mimetype)
    
    
    # "The 'iterator()' method ensures only a few rows are fetched from the database at a time, saving memory, but aren't cached if needed again. This iteractor version seems slightly faster than non-iterated version.
    # In practice, the database interface (e.g. the Python MySQLdb module) does cache data anyway. If really need to reduce memory usage, retrieve data in chunks.
    
    # Could try rawSQL: https://docs.djangoproject.com/en/dev/topics/db/sql/
    
    # Adding  only(field1, field2, ...), to this query might make a bit faster as retrieves fewer fields, so less converted to python strings/objects: https://docs.djangoproject.com/en/dev/ref/models/querysets/#only
    # dependency_list =  dependency_list.only("target_id", "wilcox_p", "effect_size", "histotype", "study_id", "interaction", "inhibitors")
    count = 0
        
    for d in dependency_list.iterator(): # was: for d in dependency_list:
        count += 1
        """
        d_json = {}
        # Using single characters for the field keys below would reduce the size of the json to be transfered to client.
        # eg: g=gene, p=wilcox_p, e=effect_size, s=study, h=histotype, i=inhibitors, a=interaction
        
        d_json['g'] = d.target_id if search_by_driver else d.driver_id
        
        # was: d_json['t'] =  d.target_id # was 'target'  # or d.target_id   ? .gene_name  # But maybe this 'id' needs to be an integer?   # Maybe don't need to ref the gene_name as the target filed is itself the gene_name
        # or on the query add:    qs.select_related('author')  # see: http://digitaldreamer.net/blog/2011/11/7/showing-foreign-key-value-django-admin-list-displa/
        # target      = models.ForeignKey(Gene, verbose_name='Target gene', db_column='target', to_field='gene_name', related_name='+', db_index=True, on_delete=models.PROTECT)
            
        d_json['p'] = format(d.wilcox_p, ".0E").replace("E-0", "E-")  # was 'wilcox_p' Scientific format and remove the leading zero from the exponent  # in template use: |stringformat:....
        d_json['e'] = d.effect_size # was 'effect_size'
        d_json['s'] = d.study_id # was: 'study_pmid' or maybe id # can use the foreign key value directly as it is the pmid. https://docs.djangoproject.com/en/1.9/topics/db/optimization/#use-foreign-key-values-directly
        # But maybe I need to call field the default name of 'study_id'
        #    study    = models.ForeignKey(Study, verbose_name='PubMed ID', db_column='pmid', to_field='pmid', on_delete=models.PROTECT, db_index=True)
            
        # + ' ' + d.full_name + ' ' + d.prevname_synonyms
            
        d_json['h'] = d.histotype   # was 'histotype' get_histotype_display()  # or could use 'd.histotype' shortened names, with a hash in javascript in the index.html file.
        d_json['i'] = '' # d.inhibitors  was 'inhibitors' - but empty for now.
        # d_json['study_summary'] = d.study.summary  # Info about all the studies separately embedded in the index.html template file is small.  dependency.study.weblink|safe
        d_json['a'] = '' # d.interaction  was: 'interaction'
        
        results.append(d_json)
        """

        """
Test stringdb:
CHECK1
RIOK2
WEE1
CDK18
TTK
PLK1
PRKCD
AURKA
GUCY2F
DCK
TWF1
BUB1
XRCC6BP1
CDKN2C
CDK11A
        """
        # As CSV, or simply each row as one array or tuple within the results array, and can optionally have a number as index, eg:
        # But cannot use gene (driver/target) as key, as dict assumes that gene is unique within this driver's data: (currently it isn't unique within histotype & study_pmid)
# was 'interaction_hhm'

        # NEED to change this target/driver search by protein ids.
        # *** should include this protein_id in related fields above for speed, as either driver or target

        interaction = d.interaction   # Medium/High/Highest. is Null? # was: 'Y' if d.interaction else '', 
        if interaction is None: interaction = ''   # should not be None, as set in table by script to ''.

        interation_protein_id = d.target.ensembl_protein_id if search_by_driver else d.driver.ensembl_protein_id
        if interation_protein_id is None: interation_protein_id = ''  # The ensembl_protein_id might be empty.
        interaction += '#'+interation_protein_id  # Always append the protein id so can use this to link to string.org
        
        
        inhibitors = d.target.inhibitors if search_by_driver else d.driver.inhibitors
        if inhibitors is None: inhibitors = ''
        
        results.append([
                    d.target_id if search_by_driver else d.driver_id, # the '_id' suffix gets the underlying gene name, rather than the foreigh key Gne object. See:  https://docs.djangoproject.com/en/1.9/topics/db/optimization/#use-foreign-key-values-directly
                    format(d.wilcox_p, ".0E").replace("E-0", "e-"),  # Scientific format, remove leading zero from the exponent
                    format(d.effect_size*100, ".1f"),  # As a percentage
                    d.histotype, # was d.get_histotype_display()  # but now using a hash in javascript to convert these shortened names.
                    d.study_id, # returns the underlying pmid number rather than the Study object
                    interaction,
                    inhibitors,  #'',  # d.inhibitors - but empty for now.
                    d.target_variant  # Just temporary to ensure display correct achilles boxplot image.
                    ])  # optionally an id: d_json['1'] = 
        """
        # If use simple CSV rather than json, then would need to check for semicolons in the input fields,
        # so is safer to use csv_writer, as it will quote any sttrings containing semicolons:
        # histotype_code and study_code return single characters to encode these fields smaller, for faster data transfer to browser.
        csv += d.target_id +div+ # the '_id' suffix gets the underlying gene name, rather than the foreigh key Gne object. See:  https://docs.djangoproject.com/en/1.9/topics/db/optimization/#use-foreign-key-values-directly
               format(d.wilcox_p, ".0E").replace("E-0", "E-") +div+  # Scientific format, remove leading zero from the exponent
               format(d.effect_size*100, ".1f") +div+  # As a percentage
               histotype_code[d.histotype] +div+ # was d.get_histotype_display()  # but now using a hash in javascript to convert these shortened names.
               study_code[d.study_id] +div+ # returns the underlying pmid number rather than the Study object
               '' +div+ # d.interaction - but empty for now
               '' +"\n" # d.inhibitors - but empty for now
        """

        # To add join to the query add:    qs.select_related('author')  # see: http://digitaldreamer.net/blog/2011/11/7/showing-foreign-key-value-django-admin-list-displa/
        # target      = models.ForeignKey(Gene, verbose_name='Target gene', db_column='target', to_field='gene_name', related_name='+', db_index=True, on_delete=models.PROTECT)
                        
        # + ' ' + d.full_name + ' ' + d.prevname_synonyms
                    
    start = get_timing(start, 'Dependency results', timing_array)
    
    # results_column_names = ['Target','Wilcox_p','Effect_size','Histotype','Study_pmid','Inhibitors','Interactions'] # Could add this to the returned 'query_info'
    histotype_details = "<b>All tissues</b>" if histotype_name == "ALL_HISTOTYPES" else ("tissue type <b>"+histotype_full_name+"</b>")
    study_details = "<b>All studies</b>" if study_pmid == "ALL_STUDIES" else ("study "+study.weblink()+" "+study.title+", "+ study.authors[:30]+" et al, "+study.journal+", "+ study.pub_date)

    query_info = {'search_by': search_by,
                  'gene_name': gene_name,
                  'gene_full_name': gene.full_name,
                  'gene_synonyms': gene.prevname_synonyms,
                  # 'gene_weblinks': gene.external_links('|'), # now passing the 'gene_ids' as a dictionary to format in webbrowser.
                  'histotype_name': histotype_name,
                  'histotype_full_name': histotype_full_name, # Not read
                  'histotype_details': histotype_details,
                  'study_pmid': study_pmid,
                  'study_details': study_details,
                  'dependency_count': count, # should be same as: dependency_list.count(), but dependency_list.count() could be another SQL query. # should be same as number of elements passed in the results array.
                  'current_url': current_url
                  }
                # study.weblink|safe
                
    print(timing_array)
    data = json.dumps({
        'success': True,
        'timings': timing_array,
        'query_info': query_info,
        'gene_ids': gene_ids_as_dictionary(gene),
        'results': results
        }, separators=[',',':']) # The default separators=[', ',': '] includes whitespace which I think would make transfer to browser larger. As ensure_ascii is True by default, the non-asciii characters are encoded as \uXXXX sequences, alternatively can set ensure_ascii to false which will allow unicode I think.
        # Alternatively use: return HttpResponse(simplejson.dumps( [drug.field for drug in drugs ]))
        # eg: format is:
        # [ {"id": "3", "value":"3","label":"Matching employee A"},
        #   {"id": "5", "value":"5","label":"Matching employee B"},
        # ]
    # data = json.dumps(list(Town.objects.filter(name__icontains=q).values('name')))
    
    start = get_timing(start, 'Json dump', timing_array) # Although too late to add this time to the json already encoded above.
    
    # BUT need to check for other process writing this file at same time - eg. include the process (os.getpid() - Returns the current process id) and maybe thread id (import _thread; _thread.get_ident() - the 'thread identifier' of the current thread. This is a nonzero integer. Its value has no direct meaning; it is intended as a magic cookie to be used e.g. to index a dictionary of thread-specific data. Thread identifiers may be recycled when a thread exits and another thread is created.)
    # then when finished writing move if safe to move to .json
    # or maybe better to store in database rather than file.
    # An example of using file system for caching - half-way down thisa webpage: https://www.pythonanywhere.com/forums/topic/197/
    #with open('Cache_'+cache_key+'.json','w') as fout:
    #    fout.write(data)
    
    cache.set(cache_key, data, version=ajax_results_cache_version) # could use the add() method instead, but better to update anyway.
    # Could gzip the cached data (using GZip middleware's gzip_page() decorator for the view, or in code https://docs.djangoproject.com/en/1.9/ref/middleware/#module-django.middleware.gzip )
    # GZipMiddleware will NOT compress content if any of the following are true:
    #  - The content body is less than 200 bytes long.
    #  - The response has already set the Content-Encoding header.
    #  - The request (the browser) hasn’t sent an Accept-Encoding header containing gzip.
    # Another option is using cache_control() permit browser caching by setting the Vary header: https://docs.djangoproject.com/en/1.9/topics/cache/#using-vary-headers
    # "(Note that the caching middleware already sets the cache header’s max-age with the value of the CACHE_MIDDLEWARE_SECONDS setting. If you use a custom max_age in a cache_control decorator, the decorator will take precedence, and the header values will be merged correctly.)"
    # https://www.pythonanywhere.com/forums/topic/376/
    # and example of gzip using flask: https://github.com/closeio/Flask-gzip
    #  https://github.com/closeio/Flask-gzip/blob/master/flask_gzip.py

    
    # Maybe better to use a JsonResponse, and use only dictionary objects (I used an array above as more comapct):
    # from django.http import JsonResponse
    # response = JsonResponse({'foo': 'bar'})
    # response.content
    # b'{"foo": "bar"}'
    # In order to serialize objects other than dict you must set the safe parameter to False:
    # response = JsonResponse([1, 2, 3], safe=False)
    # ** Before the 5th edition of EcmaScript it was possible to poison the JavaScript Array constructor. For this reason, Django does not allow passing non-dict objects to the JsonResponse constructor by default. However, most modern browsers implement EcmaScript 5 which removes this attack vector. Therefore it is possible to disable this security precaution.
    # See: https://docs.djangoproject.com/en/1.9/ref/request-response/
    
    return HttpResponse(data, content_type=json_mimetype)   # can use: charset='UTF-8' instead of putting utf-8 in the content_type

    #    context = {'dependency_list': dependency_list, 'gene': gene, 'histotype': histotype_name, 'histotype_full_name': histotype_full_name, 'study': study, 'gene_weblinks': gene_weblinks, 'current_url': current_url}
    # return render(request, 'gendep/ajax_results.html', context, content_type=mimetype) #  ??? .. charset=utf-8"


    
def stringdb_interactions(required_score, protein_list):
    print("In: get_stringdb_interactions")
    stringdb_options="network_flavor=confidence&limit=0&required_score="+required_score;  # &additional_network_nodes=0
    # The online interactive stringdb uses: "required_score" 400 and: "limit" 0 (otherwise by default will add 10 more proteins)

    protein_list = protein_list.replace(';', '%0D')  # To send to stringdb

    url = "http://string-db.org/api/psi-mi-tab/interactionsList?"+stringdb_options+"&identifiers="+protein_list;
    #print(url)
    from urllib.request import Request, urlopen
    from urllib.error import  URLError
# or maybe use streaming: http://stackoverflow.com/questions/16870648/python-read-website-data-line-by-line-when-available
# import requests
# r = requests.get(url, stream=True)
# for line in r.iter_lines():
#   if line: print line

    req = Request(url)
    try:
        response = urlopen(req)
    except URLError as e:
        if hasattr(e, 'reason'):
            return False, 'We failed to reach a server: ' + e.reason
        elif hasattr(e, 'code'):
            return False, 'The server couldn\'t fulfill the request. Error code:' + e.code
    else:  # everything is fine
        return True, response


def get_stringdb_interactions(request, required_score, protein_list):

    success, response = stringdb_interactions(required_score, protein_list) # Fetches list of actual interactions
    
    if success:
        protein_dict =  dict((protein,True) for protein in protein_list.split(';')) # Dict to check later if returned protein was in original list
    
        protein_dict2 = dict()
        for line in response:
          # if line:
            cols = line.decode('utf-8').rstrip().split("\t")
            if len(cols)<2: print("Num cols = %d: '"+line+"'" %(len(cols)))
            protein = cols[0].replace('string:', '') # as ids start with 'string:'
            if protein in protein_dict: protein_dict2[protein] = True
            else: print("*** Protein returned '%' is not in original list ***" %(protein))

            protein = cols[1].replace('string:', '') # as ids start with 'string:'
            if protein in protein_dict: protein_dict2[protein] = True
            else: print("*** Protein returned '%' is not in original list ***" %(protein))

        protein_list2 = ';'.join(protein_dict2.keys())
            
        #data = response.read().decode('utf-8') # Maybe don't need to decode it?
        # lines = data.split("\n")  # string:9606.ENSP00000302530	string:9606.ENSP00000300093
        # result = ""
        #for line in lines:
        #    print(line)
        #    cols = line.split("\t")
        #    if len(cols)<2: continue # As eg. has a newline at end so empty line at end
        #    result += cols[0].replace('string:', '')+"\t"+cols[1].replace('string:', '')+"\n"
        #print(protein_list2)
        return HttpResponse(protein_list2, content_type='text/plain') # or really: 'text/tab-separated-values', content_type=json_mimetype BUT this is tsv data
    else:
        print(response)
        return HttpResponse('ERROR: '+response, content_type='text/plain')
    # Maybe handle any exception - eg. server doesn't respond


    
def cytoscape(request, required_score, protein_list):
    success, response = stringdb_interactions(required_score, protein_list) # Fetches list of actual interactions
    
    if success:
        initial_nodes =  dict((protein,True) for protein in protein_list.split(';')) # Dict to check later if returned protein was in original list
    
        nodes = dict() # The protein nodes for cytoscape
        edges = dict()   # The edges for cytoscape
        for line in response:
          # if line:
            cols = line.decode('utf-8').rstrip().split("\t")
            if len(cols)<2: print("Num cols = %d: '"+line+"'" %(len(cols)))
            protein1 = cols[0].replace('string:', '') # as ids start with 'string:'
            if protein1 in initial_nodes:
                protein1 = protein1.replace('9606.', '')            
                nodes[protein1] = True
            else: print("*** Protein1 returned '%s' is not in original list ***" %(protein1))

            protein2 = cols[1].replace('string:', '') # as ids start with 'string:'
            if protein2 in initial_nodes:
                protein2 = protein2.replace('9606.', '')
                nodes[protein2] = True
            else: print("*** Protein2 returned '%s' is not in original list ***" %(protein2))

            edge = protein1+'#'+protein2
            edge_reversed = protein2+'#'+protein1
            if edge not in edges and edge_reversed not in edges:
                edges[edge] = True

        node_list = sorted(nodes)
        #print(node_list)
        
        edge_list = []
        for edge in edges:
            edge_list.append(edge.split('#')) # So should be array of arrays.
        # print(edge_list)

        context = {'node_list': node_list, 'edge_list': edge_list}
        return render(request, 'gendep/cytoscape.html', context)

        # return HttpResponse(protein_list2, content_type='text/plain') # or really: 'text/tab-separated-values', content_type=json_mimetype BUT this is tsv data 
    else:
        print(response)
        return HttpResponse('ERROR: '+response, content_type='text/plain')
    # Maybe handle any exception - eg. server doesn't respond

    
    
def qtip(tip):
    # Returns the ajax request to qtips
    return 
    
def gene_info(request, gene_name):
    try:
        gene = Gene.objects.get(gene_name=gene_name)
        data = { 'success': True, 'gene_name': gene.gene_name, 'full_name': gene.full_name, 'synonyms': gene.prevname_synonyms, 'ids': gene_ids_as_dictionary(gene) }  # 
    except ObjectDoesNotExist: # Not found by the objects.get()
        data = {"success": False, 'full_name': "Gene '%s' NOT found in Gene table"%(gene_name), 'message': "Gene '%s' NOT found in Gene table" %(gene_name)}
    return HttpResponse(json.dumps(data, separators=[',',':']), json_mimetype)
        
    
def graph(request, target_id):
    requested_target = get_object_or_404(Dependency, pk=target_id)
    return render(request, 'gendep/graph.html', {'target': requested_target})

def show_study(request, study_pmid):
    requested_study = get_object_or_404(Study, pk=study_pmid)
    # requested_study = get_object_or_404(Study, pk='Pending001') # Temportary for now.
    return render(request, 'gendep/study.html', {'study': requested_study})

def about(request):
    return render(request, 'gendep/about.html')

def drivers(request):
    driver_list = Gene.objects.filter(is_driver=True).order_by('gene_name')  # Needs: (is_driver=True), not just: (is_driver)
    context = {'driver_list': driver_list}
    return render(request, 'gendep/drivers.html', context)

def targets(request):
    target_list = Gene.objects.filter(is_target=True).order_by('gene_name')  # Needs: (is_driver=True), not just: (is_target)
    context = {'target_list': target_list}
    return render(request, 'gendep/targets.html', context)
    
def tissues(request):
    histotype_list = Dependency.HISTOTYPE_CHOICES
    context = {'histotype_list': histotype_list}
    return render(request, 'gendep/tissues.html', context)
    
def studies(request):
    study_list = Study.objects.order_by('pmid')
    context = {'study_list': study_list}
    return render(request, 'gendep/studies.html', context)

def faq(request):
    current_url =  request.META['HTTP_HOST'] # See download_dependencies_as_csv_file() for other, maybe better, ways to obtain currrent_url
    context = {'current_url': current_url}
    return render(request, 'gendep/faq.html', context)
    
def contact(request):
    return render(request, 'gendep/contact.html')



search_by_driver_column_headings_for_download = ['Dependency', 'Dependency description', 'Entez_id',  'Ensembl_id', 'Ensembl_protein_id', 'Dependency synonyms', 'Wilcox P-value', 'Effect size', 'Tissue', 'Inhibitors', 'String interaction', 'Study', 'PubMed Id', 'Experiment Type', 'Boxplot link']                                 

search_by_target_column_headings_for_download = ['Driver', 'Driver description', 'Entez_id',  'Ensembl_id', 'Ensembl_protein_id', 'Driver synonyms', 'Wilcox P-value', 'Effect size', 'Tissue', 'Inhibitors', 'String interaction', 'Study', 'PubMed Id', 'Experiment Type', 'Boxplot link']


# ===========================================    
def download_dependencies_as_csv_file(request, search_by, gene_name, histotype_name, study_pmid, delim_type='csv'):
    # Creates then downloads the current dependency result table as a tab-delimited file.
    # The download get link needs to contain serach_by, gene, tissue, study parameters.

    # In Windows at least, 'csv' files are associated with Excel. To also associate tsv file with excel: In your browser, create a helper preference associating file type 'text/tab-separated values' and file extensions 'tsv' with application 'Excel'. Pressing Download will then launch Excel with the data.
    import csv, time
    
    mimetype = 'text/html' # was: 'application/json'
    
    # see: http://stackoverflow.com/questions/6587393/resource-interpreted-as-document-but-transferred-with-mime-type-application-zip
    
    # For downloading large csv files, can use streaming: https://docs.djangoproject.com/en/1.9/howto/outputting-csv/#streaming-large-csv-files
    
    # request_method = request.method # 'POST' or 'GET'
    # if request_method != 'GET': return HttpResponse('Expected a GET request, but got a %s request' %(request_method), mimetype)
    # search_by = request.GET.get('search_by', "")  # It's an ajax POST request, rather than the usual ajax GET request
    # gene_name = request.GET.get('gene', "")
    # histotype_name = request.GET.get('histotype', "ALL_HISTOTYPES")
    # study_pmid = request.GET.get('study', "ALL_STUDIES")

    search_by_driver = is_search_by_driver(search_by) # Checks is valid and returns true if search_by='driver'
    
    if search_by_driver:
        select_related = [ 'target__inhibitors', search_by, 'study' ]  #  'target__ensembl_protein_id',
    else:
        select_related = [ 'driver__inhibitors', search_by, 'study' ]  # 'driver__ensembl_protein_id'
    
    error_msg, dependency_list, gene, histotype_full_name, study = build_dependency_query(search_by, gene_name, histotype_name, study_pmid, select_related=select_related, order_by='wilcox_p' ) # select_related (set to 'driver' or 'target') will include all the Gene info for the target/driver in one SQL join query, rather than doing multiple subqueries later.
    
    if error_msg != '': return HttpResponse("Error: "+error_msg, mimetype)

    #gene_weblinks = gene.external_links('|')
    # was: study__pmid=request.POST.get('study')
    # [:20] use: 'target__gene_name' instead of 'target.gene_name'
    # dependency_list_count = dependency_list.count()+1

    timestamp = time.strftime("%d-%b-%Y") # To add time use: "%H:%M:%S")

    dest_filename = ('dependency_%s_%s_%s_%s_%s.csv' %(search_by,gene_name,histotype_name,study_pmid,timestamp)).replace(' ','_') # To also replace any spaces with '_' NOTE: Is .csv as Windows will then know to open Excel, whereas if is tsv then won't

    # current_url = request.get_full_path() # To display the host in title for developing on lcalhost or pythonanywhere server.
    # current_url = request.build_absolute_uri()
    #current_url =  request.META['SERVER_NAME']
    current_url =  request.META['HTTP_HOST']
    
    # Or better method: http://stackoverflow.com/questions/17866114/django-get-absolute-url
    # from django.core.urlresolvers import reverse
    # url = request.build_absolute_uri(reverse('blog:detail', args=[blog.slug]))
    # BUT I want the static url...
    
    # Maybe use:  StaticFileStorage.url which effectively just prepends STATIC_URL to the path given,
    # but in more complex setups (eg. on S3) it does more.  ..... Another great example of when you absolutely must use the static tag: if you’re using Django’s CachedStaticFilesStorage as described in my post about Setting up static file caching, because it uses a hash of the file’s contents as part of the file name (so css/styles.css might be saved as css/styles.55e7cbb9ba48.css) and no amount of carefully constructing STATIC_URL will help you here!
    # from http://staticfiles.productiondjango.com/blog/stop-using-static-url-in-templates/
    # Setting up staic file caching: http://staticfiles.productiondjango.com/blog/setting-up-static-file-caching/
    
    
    # or create and register a 'absurl' tag for templates, which uses the above function ...
    #or request.get_host() but might not work if site behind proxy 
    # From: https://docs.djangoproject.com/en/1.9/howto/outputting-csv/
    # Using the Sites framework might be better: https://docs.djangoproject.com/en/dev/ref/contrib/sites/#how-django-uses-the-sites-framework
    # Simplest method is just to set the static url accordingly in the settings file: http://stackoverflow.com/questions/16573324/using-relative-vs-absolute-url-for-static-url-in-django
    # set DEVELOPMENT to True or False, then:
    # if DEVELOPMENT == True: STATIC_URL = '/static/'
    # else: STATIC_URL = 'https://www.mywebsite.com/static/'

    if delim_type=='csv':
        dialect = csv.excel
        content_type='text/csv' # can be called: 'application/x-csv' or 'application/csv'
    elif delim_type=='tsv':
        dialect = csv.excel_tab
        content_type='text/tab-separated-values'
    else:
        return HttpResponse("Error: Invalid delim_type='%s', as must be 'csv' or 'tsv'"%(delim_type), mimetype)

    # Create the HttpResponse object with the CSV/TSV header and downloaded filename:
    response = HttpResponse(content_type=content_type) # Maybe use the  type for tsv files?    
    response['Content-Disposition'] = 'attachment; filename="%s"' %(dest_filename)
    
    writer = csv.writer(response, dialect=dialect)  # An alternative would be: csv.unix_dialect
      # csv.excel_tab doesn't display well
      # Maybe: newline='', Can add:  quoting=csv.QUOTE_MINIMAL, or csv.QUOTE_NONE,  Dialect.delimiter,  Dialect.lineterminator
    
    count = dependency_list.count()
    study_name = "All studies" if study_pmid=='ALL_STUDIES' else study.short_name
    writer.writerows([
        ["","A total of %d dependencies were found for: %s='%s', Tissue='%s', Study='%s'" % (count, search_by.title(), gene_name,histotype_full_name, study_name),], # Added some dummy columns so Excel knows from first row that is CSV
        ["","Downloaded from CGDD on %s" %(timestamp),],
        ["",],
        ]) # Note needs the comma inside each square bracket to make python interpret each line as list than that string

    writer.writerow(search_by_driver_column_headings_for_download if search_by_driver
               else search_by_target_column_headings_for_download) # The writeheader() with 'fieldnames=' parameter is only for the DictWriter object. 

    # inhibitors =  if search_by_driver else  ....
    #    if inhibitors is None: inhibitors = ''
               
    # If could use 'target AS gene' or 'driver AS gene' in the django query then would need only one output:

    if search_by_driver: # search_by='driver'
      for d in dependency_list:  # Not using iteractor() as count() above will already have run query.
         writer.writerow([
            d.target.gene_name, # or d.target_id
            d.target.full_name, d.target.entrez_id, d.target.ensembl_id, d.target.ensembl_protein_id,
            d.target.prevname_synonyms, 
            d.wilcox_p, d.effect_size,   # wilcox_p was stringformat:".0E",  		    
            d.get_histotype_display(),
            d.target.inhibitors,  #'', # d.inhibitors,  if search_by_driver else ...
            'Yes' if d.interaction else '',  # In future this will change to Medium/High/Highest
            d.study.short_name,  d.study.pmid,  d.study.experiment_type,
            # ADD THE FULL STATIC PATH TO THE url = .... BELOW, this 'current_url' is a temporary fix: (or use: StaticFileStorage.url )
            'http://'+current_url+'/static/gendep/boxplots/'+d.boxplot_filename()  # The full url path so can paste into browser.
            ])

    else: # search_by='target'
      for d in dependency_list:  # Not using iteractor() as count() above will already have run query.
         writer.writerow([
            d.driver.gene_name,  # or d.driver_id
            d.driver.full_name, d.driver.entrez_id, d.driver.ensembl_id, d.driver.ensembl_protein_id,
            d.driver.prevname_synonyms, 
            d.wilcox_p, d.effect_size,  		    
            d.get_histotype_display(),
            d.driver.inhibitors,   # '', # d.inhibitors,  if search_by_driver else ....
            'Yes' if d.interaction else '',  # In future this will change to Medium/High/Highest
            d.study.short_name,  d.study.pmid,  d.study.experiment_type,
            # ADD THE FULL STATIC PATH TO THE url = .... BELOW, this 'current_url' is a temporary fix: (or use: StaticFileStorage.url )
            'http://'+current_url+'/static/gendep/boxplots/'+d.boxplot_filename()   # Using the full url path so can paste into browser.
            ])

    return response
    



#dialect = csv.Sniffer().sniff(csvfile.read(1024))
#    csvfile.seek(0)
#    reader = csv.reader(csvfile, dialect)
#    The excel_tab class defines the usual properties of an Excel-generated TAB-delimited file. It is registered with the dialect name 'excel-tab'.

#Dialect.delimiter    A one-character string used to separate fields. It defaults to ','.





"""
*** Finish this: download_dependencies_as_excel_file
### An advantage of Excel format is if import tsv file Excel changes eg. MARCH10 or SEP4 to a date, whereas creating the Excle fiile doesn't
# Also can add formatting, better url links, and include box-plot images.

def show_excel(request):
   # use a StringIO buffer rather than opening a file
   output = StringIO()
   w = csv.writer(output)
   for i in range(10):
      w.writerow(range(10))
   # rewind the virtual file
   output.seek(0)
   return HttpResponse(output.read(),
      )
  
  
def download_dependencies_as_excel_file(request):
    ## **** GOOD example: http://assist-software.net/blog/how-export-excel-files-python-django-application
    # http://www.developer.com/tech/article.php/10923_3727616_2/Creating-Excel-Files-with-Python-and-Django.htm
    import xlsxwriter  # need to install this python module
    # http://xlsxwriter.readthedocs.org/en/latest/index.html
    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename=%s.xlsx' %(filename)
    # xlsx_data = WriteToExcel(weather_period, town)
    # using xlsxwriter: http://xlsxwriter.readthedocs.org/worksheet.html

    # Another good alternaive is OpenPyXL: https://openpyxl.readthedocs.org/en/2.5/tutorial.html
    # eg: http://djangotricks.blogspot.co.uk/2013/12/how-to-export-data-as-excel.html
    # "Office Open XML Format - XLSX (a.k.a. OOXML or OpenXML) is a zipped, XML-based file format developed by Microsoft. It is fully supported by Microsoft Office 2007 and newer versions....
    
    # In python can add short descriptions to functions I think, eg: export_xlsx.short_description = u"Export XLSX"
    
    # or pyExcelerator: http://www.developer.com/tech/article.php/10923_3727616_2/Creating-Excel-Files-with-Python-and-Django.htm
    #  https://sourceforge.net/projects/pyexcelerator/ but last update is 2013
    # As well as setting content_type, should this function (but probably is same effect):  return HttpResponse(output.read(), mimetype='application/ms-excel')
    - answer is in Django >= 1.5 use content_type :
    import io

from django.http.response import HttpResponse

from xlsxwriter.workbook import Workbook


def your_view(request):

    output = io.BytesIO()

    workbook = Workbook(output, {'in_memory': True})
    worksheet = workbook.add_worksheet()
    worksheet.write(0, 0, 'Hello, world!')
    workbook.close()

    output.seek(0)

    response = HttpResponse(output.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response['Content-Disposition'] = "attachment; filename=test.xlsx"

    return response
    
    
    
    # or xlwt: https://pypi.python.org/pypi/xlwt  eg: http://stackoverflow.com/questions/20040965/how-to-export-data-in-python-with-excel-format
    # Create an new Excel file and add a worksheet.
    # workbook = xlsxwriter.Workbook('demo.xlsx')
    
    ### **** GOOD Example:  http://xlsxwriter.readthedocs.org/example_http_server3.html?highlight=django
    output = StringIO()
    maybe use BytesIO in Python 3??? - Most uses of StringIO need to be migrated to BytesIO, when used as a byte buffer. If used as a string buffer, the StringIO uses need to be left at StringIO. http://dafoster.net/articles/2013/04/09/making-an-existing-python-program-unicode-aware/
    
    **** GOOD: http://stackoverflow.com/questions/16393242/xlsxwriter-object-save-as-http-response-to-create-download-in-django
    
    workbook = xlsxwriter.Workbook(output)
    worksheet = workbook.add_worksheet(optional_sheet_name)

    worksheet.write('A1', 'Hello')
    workbook.close() # is needed.

    xlsx_data = output.getvalue()
   workbook.set_properties({
    'title':    'This is an example spreadsheet',
    'subject':  'With document properties',
    'author':   'John McNamara',
    'manager':  'Dr. Heinz Doofenshmirtz',
    'company':  'of Wolves',
    'category': 'Example spreadsheets',
    'keywords': 'Sample, Example, Properties',
    'comments': 'Created with Python and XlsxWriter'}) 
... also: 'status':
     'hyperlink_base':
    set_properties() - Set the document properties such as Title, Author etc.
    # see: http://xlsxwriter.readthedocs.org/workbook.html
    
    output = StringIO.StringIO()
    workbook = xlsxwriter.Workbook(output)  # as is small files, to avoid using temp files, maybe use: {'in_memory': True}
...
...
    workbook.close()    
    
 or if use with, then doesn't need close:
 with xlsxwriter.Workbook('hello_world.xlsx') as workbook:
    worksheet = workbook.add_worksheet()

 
   worksheet.write('A1', 'Hello world')
 
    # Here we will adding the code to add data
 

    xlsx_data = output.getvalue()
    
    or 
    worksheet = workbook.add_worksheet()
    
    bold = workbook.add_format({'bold': 1}) # Add a bold format to use to highlight cells.
    exponent_format = workbook.add_format({'num_format': '0.00E+00'})     # Add a number format for cells with, eg 1 x 10-4.
    
    # also can set cell colours (eg: 'bg_color') and borders, etc http://xlsxwriter.readthedocs.org/format.html

    row = 0
    ws.write_row(row, 0 column_headings_for_download, bold)
    # ['Dependency', 'Dependency description', 'Entez_id',  'Ensembl_id', 'Dependency synonyms', 'Wilcox P-value', 'Effect size', 'Tissue', 'Inhibitors', 'Known interaction', 'Study', 'PubMed Id', 'Experiment Type', 'Experiment summary']
    # ws.set_row(0, None, bold) # To make title row bold
    ws.set_column(1, 1, 20) # To make Description column (col 1) wider
    ws.set_column(4, 4, 20) # To make Synonyms column (col 1) wider

    for d in dependency_list:
        row += 1
        ws.write_string(row,  0, d.target.gene_name, bold)
        ws.write_string(row,  1, d.target.full_name)
        ws.write_string(row,  2, d.target.entrez_id)
        ws.write_string(row,  3, d.target.ensembl_id)
        ws.write_string(row,  4, d.target.prevname_synonyms)
        ws.write_number(row,  5, d.wilcox_p, exponent_format)  # |stringformat:".0E",  # was, d.wilcox_p_power10_format  but <sup>-4</sup> not that meaningful in excel
  	    ws.write_number(row,  6, d.effect_size)
        ws.write_string(row,  7, d.get_histotype_display)
        ws.write_string(row,  8, d.inhibitors)
        ws.write_boolean(row, 9, d.interaction)
        ws.write_string(row, 10, d.study.short_name)
        ws.write_url(   row, 11, url=d.study.url, string=d.study.pmid, tip='PubmedId: '+d.study.pmid+' : '+d.study.title)
        ws.write_string(row, 12, d.study.experiment_type)
        # ws.write_string(row, 13, d.study.summary)
        
        # ADD THE FULL STATIC PATH TO THE url = .... BELOW:
        ws.write_url(   row, 14, url = 'gendep/boxplots/'+d.boxplot_filename, string=d.boxplot_filename, tip='Boxplot image')

        # ws.insert_image(row, col, STATIC.....+d.boxplot_filename [, options]) # Optionally add the box-plots to excel file.
        # FileResponse objects¶


    # For large file could use:
    # FileResponse is a subclass of StreamingHttpResponse optimized for binary files. It uses wsgi.file_wrapper if provided by the wsgi server, otherwise it streams the file out in small chunks. FileResponse expects a file open in binary mode like so:
    # from django.http import FileResponse
    # response = FileResponse(open('myfile.png', 'rb'))


        
    response.write(xlsx_data)    # maybe add: mimetype='application/ms-excel'
    return response

"""
    
"""
  # Creates then downloads the currect dependency result table as an Excel file
  from openpyxl import Workbook
  from openpyxl.compat import range
  from openpyxl.cell import get_column_letter
  wb = Workbook()
  driver =
  tissue = All_tissues
  study = 'pmid'+... or All_studies
  dest_filename = 'dependency_%s_%s_%s.xlsx' %(driver,tissue,study)
  ws1 = wb.active
  ws1.title = "range names"
  for row in range(1, 40):
...     ws1.append(range(600))
  #ws2 = wb.create_sheet(title="Pi")
  # ws2['F5'] = 3.14

  for row in range(10, 20):
    for col in range(27, 54):
      _ = ws1.cell(column=col, row=row, value="%s" % get_column_letter(col))
>>> print(ws3['AA10'].value)
  # This will overwrite any existing file with the same name - so if another user tries to download this file at the same time, so maybe best to add a time stamp to file, then delay a second if file already exist, or add a suffix
  wb.save(filename = dest_filename)
"""

#def graph(request, driver_name):
#    return HttpResponse("You're looking at the graph for driver %s." % driver_name)

#def results(request, question_id):
#    response = "You're looking at the results of question %s."
#    return HttpResponse(response % question_id)

#def vote(request, question_id):
#    return HttpResponse("You're voting on question %s." % question_id)

